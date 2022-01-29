# Данная программа предназначена для анализа поддержания товарных остатков относительно матрицы
# Для корректной работы этой программы требуются следующие файлы:
#    Потери при несоблюдении 30% матрицы за отчетный месяц
#    Планирование + матрица по 25 первым позициям из первого файла
#    Ведомость по остатка (НЕкомплексная) отдельно по 4 филиалам (4 файла)
#    Файл Data1.xlsx для выгрузки полученных данных и загрузки их в дашборд

import os
import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Font, GradientFill, Alignment, PatternFill, Color
from openpyxl.writer.excel import save_workbook
import pandas as pd
import easygui
# import xlsxwriter
import copy # For copying nested dictionaries
import datetime as DT # For make a list of dates
import xlwings as xw

filename_poteri=easygui.fileopenbox()
df1 = pd.read_excel(filename_poteri, header=None)
df1.to_excel(filename_poteri + '.xlsx', index=False, header=False)
wb1 = openpyxl.load_workbook(filename_poteri + '.xlsx')
sh1 = wb1.active

# Identifying period (month, year)
dateCell=str(sh1['A5'].value)
monthNum = dateCell[5:7]
year = dateCell[8:10]
month_list={'01':'Январь', '02':'Февраль', '03':'Март', '04':'Апрель', '05':'Май', '06':'Июнь',
            '07':'Июль', '08':'Август', '09':'Сентябрь', '10':'Октябрь', '11':'Ноябрь', '12':'Декабрь'}
monthName=month_list[monthNum]
period=str(monthName+' 20'+year)

bdList = []

nameList = []
# Генератор вложенных списков
# Lists for prices
prData = [[0 for x in range(3)] for y in range(25)]
# data = [i for i in itertools.repeat(1, 5)]  Another way of generating a list

def listCor(x,y,z): # Prices list editing
    prData[x].pop(y)
    prData[x].insert(y,z)

rowP = 9
for i in range(25):
    bd = str(sh1.cell(row=rowP, column=1).value)
    name = sh1.cell(row=rowP, column=2).value
    bdList.append(bd)
    nameList.append(name)
    pr_zak = sh1.cell(row=rowP, column=10).value # Закупка
    pr_roz = sh1.cell(row=rowP, column=11).value # Розница
    nats = sh1.cell(row=rowP, column=12).value # Наценка
    # pot = sh1.cell(row=rowP, column=13).value  # may be do this one separately
    listCor(i,0,pr_zak)
    listCor(i,1,pr_roz)
    listCor(i, 2, nats)
    rowP += 1

bdPrices = dict(zip(bdList, prData))
commList = dict(zip(bdList, nameList))
wb1.save(str(filename_poteri + '.xlsx'))
wb1.close()
os.remove(filename_poteri+".xlsx")

# +++++++++++++++++++++++++++++++ Matrix file +++++++++++++++++++++++++++++++

filename_matrix=easygui.fileopenbox()
df2 = pd.read_excel(filename_matrix, header=None)
df2.to_excel(filename_matrix + 'M.xlsx', index=False, header=False)
wb2 = openpyxl.load_workbook(filename_matrix + 'M.xlsx')
sh2 = wb2.active
listLen = sh2.max_row

bdList2 = bdList[:]
bdDict = dict.fromkeys(bdList2, 0)

for i in bdList2:
    bdDict[i] = dict()
    bdDict[i]['Б33'] = []
    bdDict[i]['БД1'] = []
    bdDict[i]['БД3'] = []
    bdDict[i]['БД4'] = []

def matrixData(w,x,y,z): # Prices list editing
    bdDict[w][x].insert(y,z)

for i in bdList2:
    for j in range(1, listLen):
        if i == str(sh2.cell(row=j, column=1).value):
            m_b33 = sh2.cell(row=j, column=24).value
            m_bd1 = sh2.cell(row=j, column=38).value
            m_bd3 = sh2.cell(row=j, column=66).value
            m_bd4 = sh2.cell(row=j, column=80).value
            sr_pr_b33 = sh2.cell(row=j, column=27).value
            sr_pr_bd1 = sh2.cell(row=j, column=41).value
            sr_pr_bd3 = sh2.cell(row=j, column=69).value
            sr_pr_bd4 = sh2.cell(row=j, column=83).value
            dolya_b33 = sh2.cell(row=j, column=30).value
            dolya_bd1 = sh2.cell(row=j, column=44).value
            dolya_bd3 = sh2.cell(row=j, column=72).value
            dolya_bd4 = sh2.cell(row=j, column=86).value
            matrixData(i, 'Б33', 0, m_b33)
            matrixData(i, 'Б33', 1, sr_pr_b33)
            matrixData(i, 'Б33', 2, dolya_b33)
            matrixData(i, 'БД1', 0, m_bd1)
            matrixData(i, 'БД1', 1, sr_pr_bd1)
            matrixData(i, 'БД1', 2, dolya_bd1)
            matrixData(i, 'БД3', 0, m_bd3)
            matrixData(i, 'БД3', 1, sr_pr_bd3)
            matrixData(i, 'БД3', 2, dolya_bd3)
            matrixData(i, 'БД4', 0, m_bd4)
            matrixData(i, 'БД4', 1, sr_pr_bd4)
            matrixData(i, 'БД4', 2, dolya_bd4)

wb2.close()
os.remove(filename_matrix + 'M.xlsx')


# +++++++++++++++++++++++++++++++ B33 +++++++++++++++++++++++++++++++

filename_ved1=easygui.fileopenbox()
df3 = pd.read_excel(filename_ved1, header=None)
df3.to_excel(filename_ved1 + 'V1.xlsx', index=False, header=False)
wbv1 = openpyxl.load_workbook(filename_ved1 + 'V1.xlsx')
sh_v1 = wbv1.active
listLen_v1 = sh_v1.max_row

# Identifying period (month, year)
dateCell2=str(sh_v1['A5'].value)

startD = int(dateCell2[2:4])
startM = int(dateCell2[5:7])
startY = 2000 + int(dateCell2[8:10])
endD = int(dateCell2[14:16])
endM = int(dateCell2[17:19])
endY = 2000 + int(dateCell2[20:])

startDate = DT.datetime(startY, startM, startD)
endDate = DT.datetime(endY, endM, endD)
yeda = pd.date_range( # Creating a list of dates of the year's period
    min(startDate, endDate),
    max(startDate, endDate)
).strftime('%d.%m.%y').tolist()
yeda = tuple(yeda)

bdList3 = bdList[:]
stockDict_B33 = dict.fromkeys(bdList3, 0)

yeda_c = 0  # Counter for dates of the year
stock = 0  # Qty

for b in bdList3:  # Putting initial zeros for all bd's
    stockDict_B33[b] = {}
    for c in yeda:
        stockDict_B33[b][c] = stock

    yeda_c = 0

def yearloopB33(yc, kd, st):  # Function that puts current value
    for y in yeda[yc:]:  # till the end of the period (year)
        stockDict_B33[kd][yeda[yc]] = st
        yc += 1
    # return yc


for i in range(8, listLen_v1):
    init_stock = sh_v1.cell(row=i, column=5).value
    if init_stock == ' ':
        init_stock = 0
    bd_d = str(sh_v1.cell(row=i, column=1).value)

    if '.' not in bd_d:
        yeda_c = 0
        kod = str(sh_v1.cell(row=i, column=1).value)
        # stockDict_B33[kod] = {}
        stock = init_stock
        yearloopB33(yeda_c, kod, stock)
    else:
        stock = sh_v1.cell(row=i, column=8).value
        if stock == ' ':
            stock = 0
        date_index = yeda.index(bd_d)
        yearloopB33(date_index, kod, stock)

wbv1.close()
os.remove(filename_ved1 + 'V1.xlsx')

# +++++++++++++++++++++++++++++++ BD 1 +++++++++++++++++++++++++++++++

filename_ved2=easygui.fileopenbox()
df3 = pd.read_excel(filename_ved2, header=None)
df3.to_excel(filename_ved2 + 'V2.xlsx', index=False, header=False)
wbv2 = openpyxl.load_workbook(filename_ved2 + 'V2.xlsx')
sh_v2 = wbv2.active
listLen_v2 = sh_v2.max_row

# Identifying period (month, year)
dateCell2=str(sh_v2['A5'].value)

startD = int(dateCell2[2:4])
startM = int(dateCell2[5:7])
startY = 2000 + int(dateCell2[8:10])
endD = int(dateCell2[14:16])
endM = int(dateCell2[17:19])
endY = 2000 + int(dateCell2[20:])
"""
startDate = DT.datetime(startY, startM, startD)
endDate = DT.datetime(endY, endM, endD)
yeda = pd.date_range( # Creating a list of dates of the year's period
    min(startDate, endDate),
    max(startDate, endDate)
).strftime('%d.%m.%y').tolist()
"""
bdList4 = bdList[:]
stockDict_BD1 = dict.fromkeys(bdList4, 0)

yeda_c = 0  # Counter for dates of the year
stock = 0  # Qty

for b in bdList4:  # Putting initial zeros for all bd's
    stockDict_BD1[b] = {}
    for c in yeda:
        stockDict_BD1[b][c] = stock

yeda_c = 0

def yearloopBD1(yc, kd, st):  # Function that puts current value
    for y in yeda[yc:]:  # till the end of the period (year)
        stockDict_BD1[kd][yeda[yc]] = st
        yc += 1
    # return yc


for i in range(8, listLen_v2):
    init_stock = sh_v2.cell(row=i, column=5).value
    if init_stock == ' ':
        init_stock = 0
    bd_d = str(sh_v2.cell(row=i, column=1).value)

    if '.' not in bd_d:
        yeda_c = 0
        kod = str(sh_v2.cell(row=i, column=1).value)
        # stockDict_B33[kod] = {}
        stock = init_stock
        yearloopBD1(yeda_c, kod, stock)
    else:
        stock = sh_v2.cell(row=i, column=8).value
        if stock == ' ':
            stock = 0
        date_index = yeda.index(bd_d)
        yearloopBD1(date_index, kod, stock)


wbv2.close()
os.remove(filename_ved2 + 'V2.xlsx')

# +++++++++++++++++++++++++++++++ BD3 +++++++++++++++++++++++++++++++


filename_ved3=easygui.fileopenbox()
df3 = pd.read_excel(filename_ved3, header=None)
df3.to_excel(filename_ved3 + 'V3.xlsx', index=False, header=False)
wbv3 = openpyxl.load_workbook(filename_ved3 + 'V3.xlsx')
sh_v3 = wbv3.active
listLen_v3 = sh_v3.max_row

# Identifying period (month, year)
dateCell2=str(sh_v3['A5'].value)

startD = int(dateCell2[2:4])
startM = int(dateCell2[5:7])
startY = 2000 + int(dateCell2[8:10])
endD = int(dateCell2[14:16])
endM = int(dateCell2[17:19])
endY = 2000 + int(dateCell2[20:])
"""
startDate = DT.datetime(startY, startM, startD)
endDate = DT.datetime(endY, endM, endD)
yeda = pd.date_range( # Creating a list of dates of the year's period
    min(startDate, endDate),
    max(startDate, endDate)
).strftime('%d.%m.%y').tolist()
"""
bdList5 = bdList[:]
stockDict_BD3 = dict.fromkeys(bdList5, 0)

yeda_c = 0  # Counter for dates of the year
stock = 0  # Qty

for b in bdList5:  # Putting initial zeros for all bd's
    stockDict_BD3[b] = {}
    for c in yeda:
        stockDict_BD3[b][c] = stock

yeda_c = 0

def yearloopBD3(yc, kd, st):  # Function that puts current value
    for y in yeda[yc:]:  # till the end of the period (year)
        stockDict_BD3[kd][yeda[yc]] = st
        yc += 1
    # return yc


for i in range(8, listLen_v3):
    init_stock = sh_v3.cell(row=i, column=5).value
    if init_stock == ' ':
        init_stock = 0
    bd_d = str(sh_v3.cell(row=i, column=1).value)

    if '.' not in bd_d:
        yeda_c = 0
        kod = str(sh_v3.cell(row=i, column=1).value)
        # stockDict_B33[kod] = {}
        stock = init_stock
        yearloopBD3(yeda_c, kod, stock)
    else:
        stock = sh_v3.cell(row=i, column=8).value
        if stock == ' ':
            stock = 0
        date_index = yeda.index(bd_d)
        yearloopBD3(date_index, kod, stock)

wbv3.close()
os.remove(filename_ved3 + 'V3.xlsx')

# +++++++++++++++++++++++++++++++ BD4 ++++++++++++++++++++++++++++++++

filename_ved4=easygui.fileopenbox()
df3 = pd.read_excel(filename_ved4, header=None)
df3.to_excel(filename_ved4 + 'V4.xlsx', index=False, header=False)
wbv4 = openpyxl.load_workbook(filename_ved4 + 'V4.xlsx')
sh_v4 = wbv4.active
listLen_v4 = sh_v4.max_row

# Identifying period (month, year)
dateCell2=str(sh_v4['A5'].value)

startD = int(dateCell2[2:4])
startM = int(dateCell2[5:7])
startY = 2000 + int(dateCell2[8:10])
endD = int(dateCell2[14:16])
endM = int(dateCell2[17:19])
endY = 2000 + int(dateCell2[20:])
"""
startDate = DT.datetime(startY, startM, startD)
endDate = DT.datetime(endY, endM, endD)
yeda = pd.date_range( # Creating a list of dates of the year's period
    min(startDate, endDate),
    max(startDate, endDate)
).strftime('%d.%m.%y').tolist()
"""
bdList6 = bdList[:]
stockDict_BD4 = dict.fromkeys(bdList6, 0)

yeda_c = 0  # Counter for dates of the year
stock = 0  # Qty

for b in bdList6:  # Putting initial zeros for all bd's
    stockDict_BD4[b] = {}
    for c in yeda:
        stockDict_BD4[b][c] = stock

yeda_c = 0

def yearloopBD4(yc, kd, st):  # Function that puts current value
    for y in yeda[yc:]:  # till the end of the period (year)
        stockDict_BD4[kd][yeda[yc]] = st
        yc += 1
    # return yc


for i in range(8, listLen_v4):
    init_stock = sh_v4.cell(row=i, column=5).value
    if init_stock == ' ':
        init_stock = 0
    bd_d = str(sh_v4.cell(row=i, column=1).value)

    if '.' not in bd_d:
        yeda_c = 0
        kod = str(sh_v4.cell(row=i, column=1).value)
        # stockDict_B33[kod] = {}
        stock = init_stock
        yearloopBD4(yeda_c, kod, stock)
    else:
        stock = sh_v4.cell(row=i, column=8).value
        if stock == ' ':
            stock = 0
        date_index = yeda.index(bd_d)
        yearloopBD4(date_index, kod, stock)

wbv4.close()
os.remove(filename_ved4 + 'V4.xlsx')


# ============================== Exporting ========================

# wbm = xw.Book('Дашборд Матрица 2.xlsx')
# sh_dash = wbm.sheets['Лист1']

dashboardFile = "Data1.xlsx"
wbm = openpyxl.load_workbook(dashboardFile)
sh_dash = wbm.active
# dashboardFile = open('Дашборд Матрица 22.xlsx', 'a')

# Making a list of sheets in the dashboard file



# Exporting data

col = 7
for i in bdList:
    r = 1
    sh_dash.cell(row=1, column=col+1).value = i
    sh_dash.cell(row=1, column=col+3).value = commList[i]
    sh_dash.cell(row=6, column=col).value = bdDict[i]['Б33'][2]
    sh_dash.cell(row=6, column=col+2).value = bdDict[i]['БД1'][2]
    sh_dash.cell(row=6, column=col+4).value = bdDict[i]['БД3'][2]
    sh_dash.cell(row=6, column=col+6).value = bdDict[i]['БД4'][2]
    col += 10


col = 7
for i in bdList:
    row = 10
    for j in yeda:
        sh_dash.cell(row=row, column=1).value = j[:6] + '20' + j[6:]

        if stockDict_B33[i][j] == None:
            st = 0
        else:
            st = stockDict_B33[i][j]
        sh_dash.cell(row=row, column=col).value = st
        sh_dash.cell(row=row, column=col+1).value = bdDict[i]['Б33'][0]

        if stockDict_BD1[i][j] == None:
            st = 0
        else:
            st = stockDict_BD1[i][j]
        sh_dash.cell(row=row, column=col+2).value = st
        sh_dash.cell(row=row, column=col+3).value = bdDict[i]['БД1'][0]

        if stockDict_BD3[i][j] == None:
            st = 0
        else:
            st = stockDict_BD3[i][j]
        sh_dash.cell(row=row, column=col+4).value = st
        sh_dash.cell(row=row, column=col+5).value = bdDict[i]['БД3'][0]

        if stockDict_BD4[i][j] == None:
            st = 0
        else:
            st = stockDict_BD4[i][j]
        sh_dash.cell(row=row, column=col+6).value = st
        sh_dash.cell(row=row, column=col+7).value = bdDict[i]['БД4'][0]
        row += 1
    col += 10

wbm.save('Data1.xlsx')
wbm.close()
